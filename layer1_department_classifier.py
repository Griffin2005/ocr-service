import os
import json
import pytesseract
import concurrent.futures
from pdf2image import convert_from_path
from docx import Document
import openpyxl

# ---- CONFIG ----
INPUT_FOLDER = "input_docs"
OUTPUT_JSON = "layer1_results.json"

# Departments & keywords
DEPARTMENTS = {
    "Engineering": ["design", "engineering", "drawing", "train", "maintenance"],
    "Procurement": ["invoice", "purchase", "vendor", "tender", "contract"],
    "HR": ["policy", "leave", "recruitment", "training", "employee"],
    "Legal": ["legal", "court", "law", "compliance", "agreement"],
    "Safety": ["safety", "incident", "risk", "hazard", "security"],
    "Finance": ["budget", "finance", "accounts", "audit", "bill", "payment"]
}

# ---- FILE TYPE DETECTOR ----
def detect_file_type(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".pdf":
        return "pdf"
    elif ext in [".jpg", ".jpeg", ".png"]:
        return "image"
    elif ext in [".docx", ".doc"]:
        return "word"
    elif ext in [".xls", ".xlsx"]:
        return "excel"
    elif ext == ".txt":
        return "text"
    else:
        return "unknown"

# ---- TEXT EXTRACTORS ----
def extract_text(file_path):
    file_type = detect_file_type(file_path)
    text = ""

    try:
        if file_type == "pdf":
            try:
                from PyPDF2 import PdfReader
                reader = PdfReader(file_path)
                text = " ".join([page.extract_text() or "" for page in reader.pages])
                if not text.strip():  # fallback to OCR if empty
                    pages = convert_from_path(file_path, dpi=150, first_page=1, last_page=1)
                    text = pytesseract.image_to_string(pages[0])
            except:
                text = ""

        elif file_type == "image":
            text = pytesseract.image_to_string(file_path)

        elif file_type == "word":
            doc = Document(file_path)
            text = "\n".join([p.text for p in doc.paragraphs])

        elif file_type == "excel":
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    text += " ".join([str(cell) for cell in row if cell]) + "\n"

        elif file_type == "text":
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()

    except Exception as e:
        print(f"Error reading {file_path}: {e}")

    return text

# ---- DEPARTMENT CLASSIFICATION ----
def classify_department(text):
    scores = {dept: 0 for dept in DEPARTMENTS}
    for dept, keywords in DEPARTMENTS.items():
        for kw in keywords:
            scores[dept] += text.lower().count(kw)
    return max(scores, key=scores.get)

# ---- PROCESS ONE FILE ----
def process_file(file_path):
    text = extract_text(file_path)
    department = classify_department(text) if text.strip() else "Unknown"
    return {
        "file": os.path.basename(file_path),
        "department": department,
        "source": "Central Email/WhatsApp",
        "pages_scanned": 1,  # only first page for quick routing
        "priority": "Medium"  # can be set dynamically later
    }

# ---- MAIN MULTITHREADED RUN ----
def main():
    files = [os.path.join(INPUT_FOLDER, f) for f in os.listdir(INPUT_FOLDER)]
    results = []

    with concurrent.futures.ThreadPoolExecutor() as executor:
        future_to_file = {executor.submit(process_file, f): f for f in files}
        for future in concurrent.futures.as_completed(future_to_file):
            results.append(future.result())
            print(f"Processed: {future_to_file[future]}")

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=4)

    print(f"\nLayer 1 completed. Results saved to {OUTPUT_JSON}")

if __name__ == "__main__":
    main()
